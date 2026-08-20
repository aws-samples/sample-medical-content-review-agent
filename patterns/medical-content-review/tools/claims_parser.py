# Copyright Amazon.com, Inc. or its affiliates. All Rights Reserved.
# SPDX-License-Identifier: MIT-0
"""Pure parser for a pre-approved claims spreadsheet (.xlsx/.xlsm/.csv).

This module deliberately has no AWS or Strands dependency: it is imported both by
the `load_claims_library` agent tool and by the upload API Lambda, which parses the
spreadsheet the moment it is uploaded so the UI can show it before the review runs.
Two parsers would eventually disagree, and the preview would then show something
other than what the matcher used, so there is exactly one.
"""

from __future__ import annotations

import csv
import datetime as dt
import io
import re
from collections import Counter
from pathlib import PurePosixPath

# Canonical field -> accepted spellings of the column header (lower-cased, with
# spaces/punctuation collapsed to underscores). Anything unrecognised is kept under
# "extra" so no customer column is silently lost.
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "claim_id": ("claim_id", "claimid", "id", "claim_number", "claim_ref"),
    "claim_text": (
        "claim_text",
        "claim",
        "text",
        "approved_claim",
        "approved_text",
        "wording",
        "statement",
    ),
    "claim_type": ("claim_type", "type", "category", "claim_category"),
    "status": ("status", "approval_status", "state"),
    "approved_date": ("approved_date", "approval_date", "date_approved"),
    "expiry_date": ("expiry_date", "expiration_date", "valid_until", "review_by"),
    "reference": (
        "primary_reference",
        "reference",
        "substantiation",
        "evidence",
        "support",
    ),
    "source": ("source_document", "source", "document", "reference_document"),
    "audience": ("audience", "target_audience", "channel", "channels"),
    "restrictions": (
        "usage_restrictions",
        "restrictions",
        "conditions",
        "notes",
        "comments",
    ),
    "job_code": ("mlr_job_code", "job_code", "approval_code"),
}

CANONICAL_BY_ALIAS = {
    alias: canonical
    for canonical, aliases in COLUMN_ALIASES.items()
    for alias in aliases
}

# Header spellings the alias table cannot enumerate, resolved by pattern instead.
# Ordered most specific first, and each canonical field is filled at most once, so
# "Claim ID" can never be mistaken for the claim text column. Only the fields whose
# absence changes the outcome of the review get a fallback:
#   claim_text  — the one required column; without it nothing parses at all
#   status      — a missed status column makes a withdrawn claim look approved
#   expiry_date — same, for claims that have lapsed
#   claim_id    — only affects the id shown next to a match
_HEADER_PATTERNS: tuple[tuple[str, re.Pattern[str]], ...] = (
    (
        "claim_text",
        re.compile(
            r"claim.*(text|wording|statement|copy|message|language|description)"
            r"|(text|wording|statement|copy|message|language|description).*claim"
        ),
    ),
    ("status", re.compile(r"(^|_)(status|state)($|_)|approval.*(status|state)")),
    ("expiry_date", re.compile(r"expir|valid_(until|to|through)|lapse")),
    ("approved_date", re.compile(r"approv.*date|date.*approv")),
    ("claim_id", re.compile(r"claim.*(id|ref|code|number)|(^|_)(id|ref)($|_)")),
    # Last resort for the required column: a "claim"-ish header that is clearly not
    # one of the metadata fields above
    (
        "claim_text",
        re.compile(r"claim(?!.*(id|type|categ|status|state|date|code|ref|number))"),
    ),
)

# A header can sit below a title row, a logo row, or a blank spacer, so the header is
# looked for rather than assumed to be the first row
MAX_HEADER_SCAN_ROWS = 12


def normalise_header(value: object) -> str:
    """Reduce a raw header cell to a comparable key (lower snake case)

    Parameters
    ----------
    value : object
        The raw header cell as the spreadsheet reader returned it

    Returns
    -------
    str
        The comparable key, e.g. "Approved Claim (EU)" -> "approved_claim_eu"
    """
    text = re.sub(r"[^a-z0-9]+", "_", str(value or "").strip().lower())
    return text.strip("_")


def stringify(value: object) -> str:
    """Render a spreadsheet cell as a trimmed string, dates as ISO-8601

    Parameters
    ----------
    value : object
        The raw cell value, which openpyxl may hand over as a date or a float

    Returns
    -------
    str
        The cell rendered for display and comparison, empty when the cell is blank
    """
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _map_headers(row: tuple) -> dict[int, tuple[str, str | None]]:
    """Map each column index to its raw header and the canonical field it feeds

    Exact alias matches are resolved first for the whole row, so an unambiguous
    ``claim_id`` column always wins over the pattern fallbacks.

    Parameters
    ----------
    row : tuple
        The candidate header row

    Returns
    -------
    dict[int, tuple[str, str | None]]
        Column index -> (raw header text, canonical field or None when unmapped)
    """
    raw: dict[int, str] = {}
    for idx, cell in enumerate(row):
        text = stringify(cell)
        if text:
            raw[idx] = text

    mapped: dict[int, str | None] = {}
    taken: set[str] = set()
    for idx, text in raw.items():
        canonical = CANONICAL_BY_ALIAS.get(normalise_header(text))
        if canonical and canonical not in taken:
            mapped[idx] = canonical
            taken.add(canonical)
        else:
            mapped[idx] = None

    # Pattern fallbacks, for headers no alias covers
    for canonical, pattern in _HEADER_PATTERNS:
        if canonical in taken:
            continue
        for idx, text in raw.items():
            if mapped[idx] is not None:
                continue
            if pattern.search(normalise_header(text)):
                mapped[idx] = canonical
                taken.add(canonical)
                break

    return {idx: (text, mapped[idx]) for idx, text in raw.items()}


def _row_to_claim(
    header_map: dict[int, tuple[str, str | None]], row: tuple, fallback_index: int
) -> dict | None:
    """Turn one spreadsheet row into a normalised claim record, or None if empty

    Parameters
    ----------
    header_map : dict[int, tuple[str, str | None]]
        The mapping produced by :func:`_map_headers`
    row : tuple
        The data row
    fallback_index : int
        Row offset used to invent a claim id when the sheet has no id column

    Returns
    -------
    dict | None
        The claim record, or None when the row carries no claim text
    """
    record: dict[str, str] = {}
    extra: dict[str, str] = {}
    for col_idx, cell in enumerate(row):
        header = header_map.get(col_idx)
        if not header:
            continue
        raw_header, canonical = header
        value = stringify(cell)
        if not value:
            continue
        if canonical:
            record.setdefault(canonical, value)
        else:
            extra[raw_header] = value

    if not record.get("claim_text"):
        return None

    claim = {
        "claim_id": record.get("claim_id") or f"CLAIM-{fallback_index:03d}",
        "claim_text": record["claim_text"],
        "claim_type": record.get("claim_type", ""),
        "status": record.get("status", "Approved"),
        "approved_date": record.get("approved_date", ""),
        "expiry_date": record.get("expiry_date", ""),
        "reference": record.get("reference", ""),
        "source": record.get("source", ""),
        "audience": record.get("audience", ""),
        "restrictions": record.get("restrictions", ""),
        "job_code": record.get("job_code", ""),
    }
    if extra:
        claim["extra"] = extra
    return claim


def parse_rows(rows: list[tuple]) -> dict:
    """Parse a sheet's rows into claim records, finding the header row first

    The header is the row within the first :data:`MAX_HEADER_SCAN_ROWS` that maps the
    most canonical fields while still providing a claim text column. Title rows, logo
    rows, and blank spacers above the real header are therefore skipped instead of
    failing the whole sheet.

    Parameters
    ----------
    rows : list[tuple]
        Every row of one worksheet, header included

    Returns
    -------
    dict
        ``{"claims": [...], "columns": [...], "column_mapping": {...},
        "unmapped_columns": [...], "header_row": int, "headers_seen": [...]}``.
        ``claims`` is empty when the sheet is not a claims table; ``headers_seen``
        then reports the first non-empty row so the caller can say what it did read.
    """
    empty = {
        "claims": [],
        "columns": [],
        "column_mapping": {},
        "unmapped_columns": [],
        "header_row": 0,
        "headers_seen": [],
    }
    if not rows:
        return empty

    best_idx = -1
    best_map: dict[int, tuple[str, str | None]] = {}
    best_score = 0
    for idx, row in enumerate(rows[:MAX_HEADER_SCAN_ROWS]):
        header_map = _map_headers(row)
        canonicals = {c for _, c in header_map.values() if c}
        if "claim_text" not in canonicals:
            continue
        if len(canonicals) > best_score:
            best_idx, best_map, best_score = idx, header_map, len(canonicals)

    if best_idx < 0:
        first = next((row for row in rows if any(stringify(c) for c in row)), ())
        empty["headers_seen"] = [stringify(c) for c in first if stringify(c)]
        return empty

    claims: list[dict] = []
    for offset, row in enumerate(rows[best_idx + 1 :], start=1):
        claim = _row_to_claim(best_map, row, offset)
        if claim:
            claims.append(claim)

    mapping = {canonical: raw for raw, canonical in best_map.values() if canonical}
    return {
        "claims": claims,
        "columns": sorted(mapping),
        "column_mapping": mapping,
        "unmapped_columns": [
            raw for raw, canonical in best_map.values() if not canonical
        ],
        "header_row": best_idx + 1,
        "headers_seen": [raw for raw, _ in best_map.values()],
    }


def parse_xlsx(body: bytes) -> dict:
    """Parse the first worksheet that actually looks like a claims table

    Parameters
    ----------
    body : bytes
        The raw workbook

    Returns
    -------
    dict
        The result of :func:`parse_rows` for the winning sheet, or the last sheet's
        (empty) result when no sheet is a claims table
    """
    from openpyxl import load_workbook

    workbook = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
    result = parse_rows([])
    try:
        for sheet in workbook.worksheets:
            rows = [tuple(row) for row in sheet.iter_rows(values_only=True)]
            parsed = parse_rows(rows)
            if parsed["claims"]:
                return parsed
            if parsed["headers_seen"] and not result["headers_seen"]:
                result = parsed
    finally:
        workbook.close()
    return result


def parse_csv(body: bytes) -> dict:
    """Parse a CSV claims export

    Parameters
    ----------
    body : bytes
        The raw file, BOM tolerated

    Returns
    -------
    dict
        The result of :func:`parse_rows`
    """
    text = body.decode("utf-8-sig", errors="replace")
    return parse_rows([tuple(row) for row in csv.reader(io.StringIO(text))])


def parse_claims_file(body: bytes, filename: str) -> dict:
    """Parse a claims spreadsheet, choosing the reader from the filename

    Parameters
    ----------
    body : bytes
        The raw uploaded file
    filename : str
        Name the file was uploaded under, or the S3 key when that is all there is

    Returns
    -------
    dict
        :func:`parse_rows` output plus ``by_status`` counts

    Raises
    ------
    ValueError
        When no row carries a claim, with the headers that were read so the mismatch
        can be corrected without guessing
    """
    suffix = PurePosixPath(filename or "").suffix.lower()
    if suffix == ".csv":
        parsed = parse_csv(body)
    elif suffix in {".xlsx", ".xlsm", ".xltx", ".xltm"}:
        parsed = parse_xlsx(body)
    else:
        # Unknown extension — try Excel first, then fall back to CSV
        try:
            parsed = parse_xlsx(body)
        except Exception:
            parsed = parse_csv(body)

    if not parsed["claims"]:
        seen = ", ".join(parsed["headers_seen"][:12]) or "(no header row found)"
        raise ValueError(
            f"No claims found in {filename or 'the uploaded file'}. Columns read:"
            f" {seen}. The file needs a header row with a claim text column (e.g."
            " `claim_text`, `claim`, `approved claim wording`) and one claim per row."
        )

    parsed["by_status"] = dict(
        Counter(claim["status"] or "Approved" for claim in parsed["claims"])
    )
    return parsed


def sanitise_stem(name: str) -> str:
    """Reduce a filename to the characters that are safe inside an S3 key

    Parameters
    ----------
    name : str
        A filename or S3 key

    Returns
    -------
    str
        The sanitised stem, never empty
    """
    stem = PurePosixPath(name).stem.strip()
    stem = re.sub(r"[^A-Za-z0-9._-]+", "_", stem).strip("._-")
    return stem or "claims_library"


def library_key(source_key: str, filename: str = "", prefix: str = "claims") -> str:
    """Compute the S3 key the parsed library is written to

    Both the upload API and the agent tool parse the same upload, so both must land on
    the same key: the second write is then simply idempotent instead of leaving two
    copies. The key also carries part of the source key, because uploads are stored
    under generated names and two users uploading their own ``claims.xlsx`` must not
    end up sharing one parsed library.

    Parameters
    ----------
    source_key : str
        S3 key of the uploaded spreadsheet, e.g. ``uploads/9f8e....xlsx``
    filename : str
        Original filename, used to keep the key readable in the console
    prefix : str
        Bucket prefix to write under

    Returns
    -------
    str
        The S3 key, e.g. ``claims/library_pre_approved_claims_9f8e7d6c.json``
    """
    name_stem = sanitise_stem(filename or source_key)
    source_stem = sanitise_stem(source_key)[:8]
    return f"{prefix}/library_{name_stem}_{source_stem}.json"
